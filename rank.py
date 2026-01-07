from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager
from openpyxl import load_workbook
from datetime import datetime
import time
import os

def is_my_store(name, my_store_name):
    """띄어쓰기 무시하고 부분 문자열 매칭으로 우리 가게인지 확인"""
    name_no_space = name.replace(" ", "")
    my_store_no_space = my_store_name.replace(" ", "")
    return my_store_no_space in name_no_space or name_no_space in my_store_no_space

def search_store_rank(driver, wait, search_text, my_store_name):
    """키워드로 검색 후 상호명 순위 찾기"""
    
    driver.switch_to.default_content()
    driver.get("https://map.naver.com/")
    time.sleep(3)
    
    # 검색창 찾기
    try:
        search_input = wait.until(
            EC.presence_of_element_located((By.CSS_SELECTOR, ".input_search"))
        )
    except:
        print("  검색창을 찾을 수 없음")
        return None, None
    
    search_input.click()
    time.sleep(0.5)
    
    # GitHub Actions에서는 pyperclip 사용 불가 -> JavaScript로 입력
    if os.environ.get('GITHUB_ACTIONS'):
        driver.execute_script("arguments[0].value = arguments[1];", search_input, search_text)
        search_input.send_keys(Keys.SPACE)  # 트리거용
        search_input.send_keys(Keys.BACKSPACE)
    else:
        # 로컬에서는 pyperclip 사용
        import pyperclip
        pyperclip.copy(search_text)
        search_input.send_keys(Keys.CONTROL, 'v')
    
    time.sleep(0.5)
    search_input.send_keys(Keys.ENTER)
    print(f"  '{search_text}' 검색 완료")
    
    time.sleep(3)
    
    # iframe 전환
    try:
        search_iframe = wait.until(
            EC.presence_of_element_located((By.CSS_SELECTOR, "iframe#searchIframe"))
        )
        driver.switch_to.frame(search_iframe)
    except:
        print("  검색 결과 없음")
        return None, None
    
    time.sleep(2)
    
    store_names = []
    seen = set()
    found_rank = None
    found_name = None
    
    # 5페이지까지 검색
    for page in range(1, 6):
        print(f"  [{page}페이지 검색 중...]")
        
        try:
            scroll_container = wait.until(
                EC.presence_of_element_located((By.CSS_SELECTOR, "#_pcmap_list_scroll_container"))
            )
        except:
            print("  검색 결과 없음")
            break
        
        driver.execute_script("arguments[0].scrollTop = 0;", scroll_container)
        time.sleep(1)
        
        scroll_position = 0
        
        while True:
            driver.execute_script(f"arguments[0].scrollTop = {scroll_position};", scroll_container)
            time.sleep(0.8)
            
            store_elements = driver.find_elements(By.CSS_SELECTOR, '#_pcmap_list_scroll_container > ul > li:not([data-laim-exp-id$="*e"]) span.YwYLL')
            
            for element in store_elements:
                try:
                    name = element.text.strip()
                    if name and name not in seen:
                        seen.add(name)
                        store_names.append(name)
                        
                        if is_my_store(name, my_store_name):
                            found_rank = len(store_names)
                            found_name = name
                            break
                except:
                    pass
            
            if found_rank:
                break
            
            scroll_position += 500
            
            max_scroll = driver.execute_script("return arguments[0].scrollHeight;", scroll_container)
            current_scroll = driver.execute_script("return arguments[0].scrollTop;", scroll_container)
            client_height = driver.execute_script("return arguments[0].clientHeight;", scroll_container)
            
            if current_scroll + client_height >= max_scroll:
                time.sleep(0.5)
                store_elements = driver.find_elements(By.CSS_SELECTOR, '#_pcmap_list_scroll_container > ul > li:not([data-laim-exp-id$="*e"]) span.YwYLL')
                for element in store_elements:
                    try:
                        name = element.text.strip()
                        if name and name not in seen:
                            seen.add(name)
                            store_names.append(name)
                            
                            if is_my_store(name, my_store_name):
                                found_rank = len(store_names)
                                found_name = name
                                break
                    except:
                        pass
                break
        
        if found_rank:
            print(f"  🎯 발견! {found_rank}위")
            break
        
        if page < 5:
            try:
                next_page_btn = driver.find_element(By.CSS_SELECTOR, f"div.zRM9F > a:nth-child({page + 2})")
                next_page_btn.click()
                time.sleep(2)
            except:
                print("  더 이상 페이지 없음")
                break
    
    return found_rank, found_name


# 메인 실행
print("=" * 60)
print("📂 keyword.xlsx 파일 읽는 중...")
print("=" * 60)

try:
    wb = load_workbook('keyword.xlsx')
    ws = wb.active
except Exception as e:
    print(f"❌ 엑셀 파일 열기 실패: {e}")
    exit(1)

# 키워드/상호명 데이터 수집 (2행부터, A열: 키워드, B열: 상호명)
keywords_data = []
for row in range(2, ws.max_row + 1):
    keyword = ws[f'A{row}'].value
    store_name = ws[f'B{row}'].value
    
    if keyword and store_name:
        keywords_data.append({
            'keyword': str(keyword).strip(),
            'store_name': str(store_name).strip()
        })

if not keywords_data:
    print("❌ 검색할 데이터가 없습니다.")
    exit(1)

print(f"✅ 총 {len(keywords_data)}개의 키워드를 검색합니다.\n")

for idx, data in enumerate(keywords_data, 1):
    print(f"  {idx}. 키워드: {data['keyword']} / 상호명: {data['store_name']}")

print("\n" + "=" * 60)

# 오늘 날짜
today = datetime.now().strftime("%Y-%m-%d")

# Chrome 드라이버 설정
options = webdriver.ChromeOptions()

# GitHub Actions 환경인지 확인
if os.environ.get('GITHUB_ACTIONS'):
    print("🤖 GitHub Actions 환경 감지 - Headless 모드 활성화")
    options.add_argument("--headless=new")
    options.add_argument("--no-sandbox")
    options.add_argument("--disable-dev-shm-usage")
    options.add_argument("--disable-gpu")
    options.add_argument("--window-size=1920,1080")
else:
    options.add_argument("--start-maximized")

options.add_argument("--disable-blink-features=AutomationControlled")
options.add_argument("user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36")
options.add_experimental_option("excludeSwitches", ["enable-automation"])
options.add_experimental_option("useAutomationExtension", False)

# 드라이버 생성
service = Service(ChromeDriverManager().install())
driver = webdriver.Chrome(service=service, options=options)

driver.execute_cdp_cmd("Page.addScriptToEvaluateOnNewDocument", {
    "source": "Object.defineProperty(navigator, 'webdriver', {get: () => undefined})"
})

wait = WebDriverWait(driver, 15)

# 결과 저장할 데이터
results = []

try:
    # 각 키워드별로 검색
    for idx, data in enumerate(keywords_data, 1):
        keyword = data['keyword']
        store_name = data['store_name']
        
        print(f"\n[{idx}/{len(keywords_data)}] 키워드: '{keyword}' / 상호명: '{store_name}'")
        print("-" * 50)
        
        rank, found_name = search_store_rank(driver, wait, keyword, store_name)
        
        if rank:
            rank_text = f"{rank}위"
            print(f"  ✅ 결과: {rank}위 (검색된 상호: {found_name})")
        else:
            rank_text = "순위없음"
            print(f"  ❌ 결과: 순위없음 (5페이지 내 미발견)")
        
        results.append({
            'keyword': keyword,
            'store_name': store_name,
            'date': today,
            'rank': rank_text
        })
    
    # 결과를 엑셀에 추가 (기존 데이터 아래에 새 행으로 추가)
    print("\n" + "=" * 60)
    print("💾 결과 저장 중...")
    
    # 현재 마지막 행 찾기 (D열 기준)
    last_row = 1
    for row in range(1, ws.max_row + 1):
        if ws[f'D{row}'].value is not None:
            last_row = row
    
    # 헤더가 없으면 추가
    if ws['D1'].value is None:
        ws['D1'] = '날짜'
        ws['E1'] = '순위'
        ws['A1'] = '키워드'
        ws['B1'] = '상호명'
    
    # 새로운 행에 결과 추가
    new_row = last_row + 1
    for result in results:
        ws[f'A{new_row}'] = result['keyword']
        ws[f'B{new_row}'] = result['store_name']
        ws[f'D{new_row}'] = result['date']
        ws[f'E{new_row}'] = result['rank']
        new_row += 1
    
    wb.save('keyword.xlsx')
    
    print(f"✅ {len(results)}개의 결과가 추가되었습니다.")
    print("📁 결과가 keyword.xlsx 파일에 저장되었습니다.")
    print("=" * 60)
    
except Exception as e:
    print(f"\n❌ 오류 발생: {e}")
    # 오류 발생 시에도 저장 시도
    try:
        wb.save('keyword.xlsx')
        print("💾 현재까지 결과 저장 완료")
    except:
        pass

finally:
    driver.quit()
    print("\n브라우저 종료")

# 로컬 실행시에만 입력 대기
if not os.environ.get('GITHUB_ACTIONS'):
    input("\n종료하려면 Enter를 누르세요...")
